"""
excel_parser.py — Excel (.xlsx) / CSV 词表解析器

输出与 pdf_parser 一致的 entries 结构：
    [{'english': str, 'chinese': str, 'phonetic': str, 'pos': str, 'failed': bool}, ...]

设计要点：
- xlsx 用 openpyxl read_only 流式读取（仅第一个 Sheet）
- csv 编码降级：utf-8-sig → utf-8 → gbk
- 列映射方案①：用户只选英文/中文列；音标/词性按列名自动识别
"""

import csv
import os
import re
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # 延迟报错，给更友好的提示


# ── 列名识别字典（小写匹配） ──────────────────────────────
ENGLISH_HEADERS = {'word', 'words', 'english', 'vocabulary', 'vocab',
                   '单词', '英文', '英语', '词汇', '文章'}
CHINESE_HEADERS = {'meaning', 'meanings', 'chinese', 'translation', 'definition',
                   'definitions', '释义', '中文', '解释', '意思', '翻译', '题目'}
PHONETIC_HEADERS = {'phonetic', 'phonetics', 'pronunciation', 'ipa',
                    '音标', '发音', '英标', '美标'}
POS_HEADERS = {'pos', 'part of speech', 'partofspeech',
               '词性', '词类'}
SYNONYM_HEADERS = {'synonym', 'synonyms', 'syn', 'syns', 'similar',
                   '同义词', '近义词', '同义', '近义'}

# CSV 编码降级序列
_CSV_ENCODINGS = ['utf-8-sig', 'utf-8', 'gbk']


# ─────────────────────────────────────────
# 文件读取（统一返回 list[list[str]]）
# ─────────────────────────────────────────

def _normalize_cell(v) -> str:
    """openpyxl 单元格可能是 None/int/float/datetime → 统一转 str"""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_xlsx(filepath: str) -> list[list[str]]:
    """读取 xlsx 第一个 Sheet 所有行"""
    if load_workbook is None:
        raise RuntimeError('未安装 openpyxl，请运行 pip install openpyxl')
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f'Excel 文件解析失败：{e}')

    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise RuntimeError('Excel 中没有任何 Sheet')
        ws = wb[sheet_names[0]]

        rows: list[list[str]] = []
        for raw_row in ws.iter_rows(values_only=True):
            cells = [_normalize_cell(c) for c in raw_row]
            # 全部为空的行 → 跳过
            if not any(cells):
                continue
            rows.append(cells)
        return rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_csv(filepath: str) -> list[list[str]]:
    """读取 CSV（编码降级）"""
    last_err = None
    for enc in _CSV_ENCODINGS:
        try:
            with open(filepath, 'r', encoding=enc, newline='') as f:
                reader = csv.reader(f)
                rows: list[list[str]] = []
                for r in reader:
                    cells = [(c or '').strip() for c in r]
                    if not any(cells):
                        continue
                    rows.append(cells)
            return rows
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except Exception as e:
            raise RuntimeError(f'CSV 文件解析失败：{e}')
    raise RuntimeError('CSV 文件编码无法识别，请用 UTF-8 保存')


def parse_table_raw(filepath: str) -> list[list[str]]:
    """根据扩展名分发读取，返回所有非空行（含可能的表头行）"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xlsx':
        rows = _read_xlsx(filepath)
    elif ext == '.csv':
        rows = _read_csv(filepath)
    else:
        raise RuntimeError(f'不支持的文件类型：{ext}')

    if not rows:
        raise RuntimeError('文件中未读到任何数据')

    # 统一列数（短行末尾补空字符串）
    max_cols = max(len(r) for r in rows)
    rows = [r + [''] * (max_cols - len(r)) for r in rows]
    return rows


# ─────────────────────────────────────────
# 智能识别
# ─────────────────────────────────────────

_CHINESE_CHAR_RE = re.compile(r'[\u4e00-\u9fa5]')
_ASCII_LETTER_RE = re.compile(r'[a-zA-Z]')


def _ratio_chinese(text: str) -> float:
    """字符串中中文字符占比（含字母 + 中文的字符基数）"""
    if not text:
        return 0.0
    chinese = len(_CHINESE_CHAR_RE.findall(text))
    ascii_l = len(_ASCII_LETTER_RE.findall(text))
    total = chinese + ascii_l
    if total == 0:
        return 0.0
    return chinese / total


def _norm_header(h: str) -> str:
    """归一化表头：去空格、转小写"""
    return re.sub(r'\s+', ' ', (h or '').strip().lower())


def looks_like_header(first_row: list[str]) -> bool:
    """启发式判定第一行是否为表头。

    规则：
    1. 第一行所有单元格都不含中文字符
    2. 且至少一个单元格匹配已知表头词（英文/中文/音标/词性/同义词同义词集）
    3. 单元格内容普遍较短（避免误把整段释义当表头）
    """
    if not first_row:
        return False
    cells = [c for c in first_row if c]
    if not cells:
        return False

    # 含中文 → 大概率不是表头
    if any(_CHINESE_CHAR_RE.search(c) for c in cells if _norm_header(c) not in
           (CHINESE_HEADERS | PHONETIC_HEADERS | POS_HEADERS | ENGLISH_HEADERS | SYNONYM_HEADERS)):
        return False

    # 至少一个单元格匹配已知表头词
    known = ENGLISH_HEADERS | CHINESE_HEADERS | PHONETIC_HEADERS | POS_HEADERS | SYNONYM_HEADERS
    if not any(_norm_header(c) in known for c in cells):
        return False

    # 普遍较短（每个 <= 30 字符）
    if any(len(c) > 30 for c in cells):
        return False

    return True


def guess_columns(rows: list[list[str]]) -> dict:
    """智能预选英文/中文/音标/词性/同义词列。

    返回 dict：
        {
            'english_col': int,       # -1 表示未识别
            'chinese_col': int,
            'phonetic_col': int,
            'pos_col': int,
            'synonym_col': int,
            'skip_first_row': bool,
            'phonetic_label': str,
            'pos_label': str,
            'synonym_label': str,
            'suggested_mode': str,    # 'standard' | 'synonym'
                                      # 释义列中文占比 < 10% 时为 'synonym'
        }
    """
    result = {
        'english_col': -1,
        'chinese_col': -1,
        'phonetic_col': -1,
        'pos_col': -1,
        'synonym_col': -1,
        'skip_first_row': False,
        'phonetic_label': '',
        'pos_label': '',
        'synonym_label': '',
        'suggested_mode': 'standard',
    }
    if not rows:
        return result

    first = rows[0]
    n_cols = len(first)
    is_header = looks_like_header(first)
    result['skip_first_row'] = is_header

    # ── Step 1: 按列名预选（仅当像表头时） ──
    if is_header:
        for ci, cell in enumerate(first):
            key = _norm_header(cell)
            if key in ENGLISH_HEADERS and result['english_col'] == -1:
                result['english_col'] = ci
            elif key in CHINESE_HEADERS and result['chinese_col'] == -1:
                result['chinese_col'] = ci
            elif key in PHONETIC_HEADERS and result['phonetic_col'] == -1:
                result['phonetic_col'] = ci
                result['phonetic_label'] = cell
            elif key in POS_HEADERS and result['pos_col'] == -1:
                result['pos_col'] = ci
                result['pos_label'] = cell
            elif key in SYNONYM_HEADERS and result['synonym_col'] == -1:
                result['synonym_col'] = ci
                result['synonym_label'] = cell

    # ── Step 2: 按内容预选剩余的英文/中文列 ──
    data_rows = rows[1:] if is_header else rows
    sample = data_rows[:10]
    if not sample:
        sample = rows[:10]

    col_chinese_ratio: list[float] = []
    col_ascii_score: list[float] = []
    for ci in range(n_cols):
        col_cells = [r[ci] for r in sample if ci < len(r) and r[ci]]
        if not col_cells:
            col_chinese_ratio.append(0.0)
            col_ascii_score.append(0.0)
            continue
        ratios = [_ratio_chinese(c) for c in col_cells]
        col_chinese_ratio.append(sum(ratios) / len(ratios))
        # ASCII score = 1 - chinese ratio
        col_ascii_score.append(1.0 - col_chinese_ratio[-1])

    # 英文列：选 ASCII 得分最高且未被占用的列
    if result['english_col'] == -1:
        best_ci, best_score = -1, -1.0
        for ci in range(n_cols):
            if ci in (result['chinese_col'], result['phonetic_col'], result['pos_col'], result['synonym_col']):
                continue
            if col_ascii_score[ci] > best_score and col_ascii_score[ci] > 0.5:
                best_score = col_ascii_score[ci]
                best_ci = ci
        result['english_col'] = best_ci

    # 中文列：优先选中文比例最高的列；若全部列中文占比都低，则退化为
    # "选 ASCII 占比最高的剩余列"作为释义列（兼容英文-英文同义词词库）
    if result['chinese_col'] == -1:
        best_ci, best_score = -1, -1.0
        for ci in range(n_cols):
            if ci in (result['english_col'], result['phonetic_col'], result['pos_col'], result['synonym_col']):
                continue
            if col_chinese_ratio[ci] > best_score and col_chinese_ratio[ci] > 0.2:
                best_score = col_chinese_ratio[ci]
                best_ci = ci
        if best_ci == -1:
            # 退化策略：所有列中文占比都 < 0.2 时，挑剩余列里 ASCII 占比最高的
            fallback_ci, fallback_score = -1, -1.0
            for ci in range(n_cols):
                if ci in (result['english_col'], result['phonetic_col'], result['pos_col'], result['synonym_col']):
                    continue
                if col_ascii_score[ci] > fallback_score and col_ascii_score[ci] > 0.5:
                    fallback_score = col_ascii_score[ci]
                    fallback_ci = ci
            best_ci = fallback_ci
        result['chinese_col'] = best_ci

    # 推断导入模式：释义列（chinese_col）的中文占比 < 10% 时推荐同义词模式
    if result['chinese_col'] >= 0:
        cn_ratio = col_chinese_ratio[result['chinese_col']] if result['chinese_col'] < len(col_chinese_ratio) else 0.0
        result['suggested_mode'] = 'synonym' if cn_ratio < 0.1 else 'standard'
    else:
        result['suggested_mode'] = 'standard'

    return result


# ─────────────────────────────────────────
# 应用映射 → 标准 entries
# ─────────────────────────────────────────

def apply_mapping(
    rows: list[list[str]],
    english_col: int,
    chinese_col: int,
    phonetic_col: int = -1,
    pos_col: int = -1,
    synonym_col: int = -1,
    skip_first_row: bool = True,
    import_mode: str = 'standard',
) -> list[dict]:
    """根据用户映射，将原始行转换为标准 entries

    import_mode:
      'standard' — B 列只写入 chinese 字段（默认行为）
      'synonym'  — B 列同时写入 chinese 和 synonyms 字段（同义词词库专用）
                   显式 synonym_col 优先于 import_mode 推断的同义词。
    """
    if english_col < 0 or chinese_col < 0:
        raise RuntimeError('必须指定英文列和中文列')
    if import_mode not in ('standard', 'synonym'):
        raise RuntimeError(f'不支持的导入模式：{import_mode}')

    data = rows[1:] if skip_first_row else rows
    entries: list[dict] = []

    def _get(row: list[str], ci: int) -> str:
        if ci < 0 or ci >= len(row):
            return ''
        return (row[ci] or '').strip()

    for row in data:
        english = _get(row, english_col)
        chinese = _get(row, chinese_col)
        phonetic = _get(row, phonetic_col)
        pos = _get(row, pos_col)
        synonyms = _get(row, synonym_col)

        # 同义词模式：若没有显式同义词列，则把 chinese 复制到 synonyms
        if import_mode == 'synonym' and not synonyms:
            synonyms = chinese

        # 整行空白 → 跳过（不计入 entries）
        if not english and not chinese and not phonetic and not pos and not synonyms:
            continue

        failed = (not english) or (not chinese)
        entries.append({
            'english': english,
            'chinese': chinese,
            'phonetic': phonetic,
            'pos': pos,
            'synonyms': synonyms,
            'failed': failed,
        })

    return entries
